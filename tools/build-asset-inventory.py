"""Build the master art-asset inventory workbook for Aetherfall.

Outputs:
  docs/asset-inventory.xlsx  — human-friendly multi-sheet workbook
  docs/asset-manifest.json   — machine-readable counterpart

Each sheet is one asset category. The columns describe what the
image-gen / sprite artist needs: file path, dimensions, format,
purpose, and a starter generation prompt where useful.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_XLSX = ROOT / "docs" / "asset-inventory.xlsx"
OUT_JSON = ROOT / "docs" / "asset-manifest.json"

CARDS_JSON = Path("/tmp/cards.json")

HEADER_FILL = PatternFill("solid", start_color="1F2A36")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="F5D083")
SECTION_FILL = PatternFill("solid", start_color="2D3A48")
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color="F5E1A8")
BODY_FONT = Font(name="Calibri", size=10)
THIN = Side(style="thin", color="3A4554")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def style_header(ws, row, columns):
    for col_index, _ in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_sheet(ws, columns, rows):
    ws.append(columns)
    style_header(ws, 1, columns)
    for r in rows:
        ws.append(r)
    for col_index, _ in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = 26
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(columns)):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    ws.freeze_panes = "A2"


# ---------- DATA SOURCES ----------

with CARDS_JSON.open() as f:
    CARDS = json.load(f)

CLASSES = [
    {"id": "rook", "name": "Rook, the Iron Vanguard", "shortName": "Rook", "role": "Defender",
     "palette": "warm bronze + ember", "banner_status": "needs new banner", "portrait_status": "needs head crop"},
    {"id": "lyra", "name": "Lyra, the Ember Veil", "shortName": "Lyra", "role": "Striker",
     "palette": "forest green + ember", "banner_status": "needs new banner", "portrait_status": "needs head crop"},
    {"id": "virex", "name": "Captain Virex, the Loot Fleet", "shortName": "Virex", "role": "Opportunist",
     "palette": "teal + brass", "banner_status": "needs new banner", "portrait_status": "needs head crop"},
    {"id": "elyra", "name": "Elyra Rootweaver, Forest Golem Architect", "shortName": "Elyra", "role": "Builder",
     "palette": "verdant + bone-bark", "banner_status": "needs new banner", "portrait_status": "needs head crop"},
    {"id": "gorath", "name": "Gorath Deepbreaker, Abyss Hunter", "shortName": "Gorath", "role": "Berserker",
     "palette": "deep violet + abyss black", "banner_status": "needs new banner", "portrait_status": "needs head crop"},
    {"id": "storm-forge", "name": "Storm Forge Vanguard", "shortName": "Storm Forge", "role": "Striker",
     "palette": "violet lightning + void black", "banner_status": "DONE — banners/storm-forge.png",
     "portrait_status": "auto-cropped from banner"},
    {"id": "verdant-reach", "name": "Verdant Reach, Greenheart Sentinel", "shortName": "Verdant Reach", "role": "Sustain",
     "palette": "moss green + bio-light", "banner_status": "needs new banner", "portrait_status": "needs head crop"},
    {"id": "tideflow-engineer", "name": "Tideflow Engineer", "shortName": "Tideflow", "role": "Support",
     "palette": "deep ocean + cyan", "banner_status": "needs new banner", "portrait_status": "needs head crop"},
    {"id": "hydroflow", "name": "Hydroflow Adept", "shortName": "Hydroflow", "role": "Support",
     "palette": "cyan water + indigo", "banner_status": "DONE — banners/hydroflow.png",
     "portrait_status": "auto-cropped from banner"},
]

MONSTERS = [
    {"id": "hollow-titan", "name": "The Hollow Titan", "role": "Boss", "encounter": "hollow-titan"},
    {"id": "ironjaw-bruiser", "name": "Ironjaw Bruiser", "role": "Frontline / Tank", "encounter": "ironjaw-bruiser"},
    {"id": "warden-of-targeting", "name": "Warden of Targeting", "role": "Surveillance Construct", "encounter": "warden-of-targeting"},
    {"id": "siege-mauler", "name": "Siege Mauler", "role": "Frontline Bruiser", "encounter": "bruiser-duo"},
    {"id": "savage-hound", "name": "Savage Hound", "role": "Aggressive Hunter", "encounter": "bruiser-duo"},
    {"id": "execution-drone", "name": "Execution Drone", "role": "Finisher", "encounter": "synthetic-hunter-squad"},
    {"id": "signal-commander", "name": "Signal Commander", "role": "Support Buffer", "encounter": "synthetic-hunter-squad"},
    {"id": "bulwark-unit", "name": "Bulwark Unit", "role": "Defensive Frontline", "encounter": "synthetic-hunter-squad"},
]

ENCOUNTERS = [
    {"id": "bruiser-duo", "name": "Bruiser Duo", "monsters": "siege-mauler, savage-hound",
     "banner_status": "DONE — banners/bruiser-duo.png"},
    {"id": "synthetic-hunter-squad", "name": "Synthetic Hunter Squad", "monsters": "signal-commander, bulwark-unit, execution-drone",
     "banner_status": "DONE — banners/synthetic-hunter-squad.png"},
    {"id": "hollow-titan", "name": "The Hollow Titan", "monsters": "hollow-titan", "banner_status": "needs banner"},
    {"id": "ironjaw-bruiser", "name": "Ironjaw Bruiser", "monsters": "ironjaw-bruiser", "banner_status": "needs banner"},
    {"id": "warden-of-targeting", "name": "Warden of Targeting", "monsters": "warden-of-targeting", "banner_status": "needs banner"},
]

TOKENS = [
    {"id": "storm-charge", "label": "Storm Charge", "color": "#c084ff (violet)",
     "passive": "+1 first damage action per turn while held",
     "icon_size": "64x64 PNG transparent",
     "prompt": "lightning bolt sigil with electric arc trails, violet-purple glow, dark void background, "
               "transparent edges, sci-fi etched-glass treatment"},
    {"id": "hydroflow", "label": "Hydroflow", "color": "#3ec4ff (cyan)",
     "passive": "Reduce 2+ threat gains by 1 while held",
     "icon_size": "64x64 PNG transparent",
     "prompt": "water droplet sigil with circular flow trails, cyan glow, deep blue void, transparent edges, "
               "sci-fi etched-glass treatment"},
    {"id": "bio-growth", "label": "Bio-Growth", "color": "#5cd66b (green)",
     "passive": "+1 to each healing action while held",
     "icon_size": "64x64 PNG transparent",
     "prompt": "leaf sigil with bio-luminescent vein trails, vibrant green glow, dark forest void, "
               "transparent edges, sci-fi etched-glass treatment"},
]

UI_ICONS = [
    {"name": "HP / Heart",      "key": "hp",      "size": "32x32",  "format": "SVG / PNG",  "color": "red ramp",  "usage": ".stat-pill.stat-hp"},
    {"name": "Energy / Bolt",   "key": "energy",  "size": "32x32",  "format": "SVG / PNG",  "color": "gold/amber","usage": ".stat-pill.stat-energy"},
    {"name": "Block / Shield-filled", "key": "block",   "size": "32x32", "format": "SVG / PNG", "color": "cyan", "usage": ".stat-pill.stat-block"},
    {"name": "Defense / Shield-outline", "key": "defense", "size": "32x32", "format": "SVG / PNG", "color": "slate", "usage": ".stat-pill.stat-def (monster)"},
    {"name": "Threat / Target", "key": "threat",  "size": "32x32",  "format": "SVG / PNG",  "color": "warm orange","usage": ".threat-pip"},
    {"name": "Fixate / Marked", "key": "fixate",  "size": "32x32",  "format": "SVG / PNG",  "color": "deep red",  "usage": "fixate ring overlay"},
    {"name": "Sword / Attack",  "key": "attack",  "size": "32x32",  "format": "SVG / PNG",  "color": "tier accent","usage": "intent chip + attack card role"},
    {"name": "Aegis / Defend",  "key": "defend",  "size": "32x32",  "format": "SVG / PNG",  "color": "cyan",      "usage": "defense card role"},
    {"name": "Ember / Heal",    "key": "heal",    "size": "32x32",  "format": "SVG / PNG",  "color": "warm green","usage": "healing card role"},
    {"name": "Banner / Support","key": "support", "size": "32x32",  "format": "SVG / PNG",  "color": "gold",      "usage": "support card role"},
    {"name": "Star / Unique",   "key": "unique",  "size": "32x32",  "format": "SVG / PNG",  "color": "lilac",     "usage": "unique card role"},
    {"name": "Pack Hunter",     "key": "pack-hunter", "size": "32x32", "format": "SVG / PNG", "color": "blood orange", "usage": "Bruiser Duo trait badge"},
    {"name": "Crossfire",       "key": "crossfire",   "size": "32x32", "format": "SVG / PNG", "color": "violet",       "usage": "Execution Drone trait badge"},
    {"name": "Target Uplink",   "key": "target-uplink", "size": "32x32", "format": "SVG / PNG", "color": "purple",     "usage": "Signal Commander trait badge"},
    {"name": "Shared Shielding","key": "shared-shielding", "size": "32x32", "format": "SVG / PNG", "color": "steel",   "usage": "Bulwark Unit trait badge"},
]

UI_FRAMES = [
    {"name": "Stat pill backplate", "key": "frame-stat-pill", "size": "1x22 (9-slice horizontal)",
     "format": "PNG with transparency", "notes": "rounded pill, faint inner highlight, used by .stat-pill"},
    {"name": "Baseplate panel",     "key": "frame-baseplate", "size": "9-slice 320x88",
     "format": "PNG with transparency", "notes": "rounded rect, subtle gradient, replaces .baseplate gradient"},
    {"name": "Action bar background", "key": "frame-action-bar", "size": "1920x76 (or 9-slice horizontal)",
     "format": "PNG with transparency", "notes": "fixed bottom bar; chamfered top corners"},
    {"name": "Threat orb",          "key": "frame-threat-orb", "size": "96x96",
     "format": "PNG with transparency", "notes": "floating round indicator that says 'X to <name>'"},
    {"name": "Battle log frame",    "key": "frame-log",       "size": "9-slice 320x420",
     "format": "PNG with transparency", "notes": "top-right corner widget; chamfered top-left + top-right"},
    {"name": "Token chip backplate","key": "frame-token-chip","size": "9-slice 80x24",
     "format": "PNG with transparency", "notes": "elemental glow, faction-tinted; used by .token-chip"},
    {"name": "Hand card frame",     "key": "frame-card",      "size": "9-slice 158x210",
     "format": "PNG with transparency", "notes": "decorative frame surrounding card art / text; per-faction palette"},
    {"name": "Intent chip frame",   "key": "frame-intent",    "size": "9-slice horizontal",
     "format": "PNG with transparency", "notes": "tier-accent border + soft inner shadow"},
    {"name": "Round summary toast", "key": "frame-summary",   "size": "9-slice 540x200",
     "format": "PNG with transparency", "notes": "centered toast above action bar"},
    {"name": "Setup encounter card","key": "frame-setup-card","size": "9-slice 600x340",
     "format": "PNG with transparency", "notes": "encounter picker hero card on the setup screen"},
]

EFFECTS = [
    {"name": "Damage hit",      "key": "fx-damage",     "size": "256x256 (9-frame strip 64x256)", "format": "PNG sprite-sheet", "notes": "8-frame sword slash + spark"},
    {"name": "Heal pulse",      "key": "fx-heal",       "size": "256x256 (8-frame strip)",        "format": "PNG sprite-sheet", "notes": "soft green wash + leaf trails"},
    {"name": "Block / shield",  "key": "fx-block",      "size": "256x256 (8-frame strip)",        "format": "PNG sprite-sheet", "notes": "hex-grid shield bloom"},
    {"name": "Token gain",      "key": "fx-token-gain", "size": "256x256 (12-frame strip)",       "format": "PNG sprite-sheet", "notes": "elemental sigil swirling in"},
    {"name": "Card play burst", "key": "fx-card-play",  "size": "512x256 (8-frame)",              "format": "PNG sprite-sheet", "notes": "horizontal sweep + faction sparks"},
    {"name": "Death / kill",    "key": "fx-death",      "size": "512x512 (16-frame)",             "format": "PNG sprite-sheet", "notes": "shatter + dissolve, monster falls"},
    {"name": "Threat gain",     "key": "fx-threat",     "size": "256x256 (8-frame)",              "format": "PNG sprite-sheet", "notes": "warm orange ember rising over portrait"},
    {"name": "Fixate lock-on",  "key": "fx-fixate",     "size": "256x256 (12-frame)",             "format": "PNG sprite-sheet", "notes": "red brackets snap onto target portrait"},
    {"name": "Spell hit",       "key": "fx-spell",      "size": "256x256 (8-frame)",              "format": "PNG sprite-sheet", "notes": "purple arc + impact rings"},
    {"name": "Critical line",   "key": "fx-critical",   "size": "256x256 (12-frame)",             "format": "PNG sprite-sheet", "notes": "yellow flash + gold motes"},
]

BACKGROUNDS = [
    {"name": "Battlefield (in-fight)",     "key": "bg-battlefield", "size": "1920x1080+ (16:9)", "format": "PNG/JPG",  "status": "DONE — assets/ui/battlefield-aetherfall.png"},
    {"name": "Title splash",               "key": "bg-title",       "size": "1920x1080 (16:9)", "format": "PNG/JPG",  "status": "TODO — see Menu/Title sheet"},
    {"name": "Title splash parallax — far","key": "bg-title-far",   "size": "2400x1200 (oversize for parallax)", "format": "PNG with transparency", "status": "TODO"},
    {"name": "Title splash parallax — mid","key": "bg-title-mid",   "size": "2400x1200", "format": "PNG with transparency", "status": "TODO"},
    {"name": "Title splash parallax — near","key": "bg-title-near", "size": "2400x1200", "format": "PNG with transparency", "status": "TODO"},
    {"name": "Setup screen ambient",       "key": "bg-setup",       "size": "1920x1080", "format": "PNG/JPG",  "status": "TODO — currently reuses battlefield"},
    {"name": "Loading transition",         "key": "bg-loading",     "size": "1920x1080", "format": "PNG/JPG",  "status": "TODO"},
    {"name": "Victory splash",             "key": "bg-victory",     "size": "1920x1080", "format": "PNG/JPG",  "status": "TODO"},
    {"name": "Defeat splash",              "key": "bg-defeat",      "size": "1920x1080", "format": "PNG/JPG",  "status": "TODO"},
]

SPRITESHEETS = [
    {"name": "All token icons",        "key": "sheet-tokens",      "layout": "3 cols x 1 row", "cell": "64x64", "total": "192x64", "contains": "storm, hydro, bio"},
    {"name": "All UI role icons",      "key": "sheet-roles",       "layout": "5 cols x 1 row", "cell": "32x32", "total": "160x32", "contains": "attack, defend, heal, support, unique"},
    {"name": "All stat icons",         "key": "sheet-stats",       "layout": "6 cols x 1 row", "cell": "32x32", "total": "192x32", "contains": "hp, energy, block, defense, threat, fixate"},
    {"name": "Monster trait badges",   "key": "sheet-traits",      "layout": "4 cols x 1 row", "cell": "32x32", "total": "128x32", "contains": "pack hunter, crossfire, target uplink, shared shielding"},
    {"name": "Effect strip — damage",  "key": "sheet-fx-damage",   "layout": "8 cols x 1 row", "cell": "64x64", "total": "512x64", "contains": "8 frames per effect"},
    {"name": "Card faction frames",    "key": "sheet-card-frames", "layout": "9 cols x 1 row", "cell": "158x210", "total": "1422x210", "contains": "one frame per playable class"},
    {"name": "Card backs",             "key": "sheet-card-backs",  "layout": "9 cols x 1 row", "cell": "158x210", "total": "1422x210", "contains": "one back per playable class"},
]

MENU_SCREENS = [
    {"screen": "Boot logo", "element": "Studio / engine logo", "notes": "Fade in 800ms hold 1500ms fade out 600ms"},
    {"screen": "Title splash", "element": "Background (parallax 3-layer)", "notes": "Mouse-tracking offset 1-3% per layer"},
    {"screen": "Title splash", "element": "Game logo / title art", "notes": "Centered, scales on hover; 1024x256 PNG transparent"},
    {"screen": "Title splash", "element": '"Press any key" prompt', "notes": "Fades in after 1.2s, soft slow breath; 480x40 PNG"},
    {"screen": "Title splash", "element": "Ambient particles (embers / dust)", "notes": "tsParticles or canvas — 80-120 particles, mouse-repulse"},
    {"screen": "Title splash", "element": "Atmospheric audio loop", "notes": "Background drone music, 60-120s loop, ducks 60% on key press"},
    {"screen": "Main menu", "element": "Menu items (Continue / New / Options / Credits)", "notes": "Hover scale 1.06 + glow + class-tinted underline"},
    {"screen": "Encounter selection", "element": "Encounter cards (3D tilt on hover)", "notes": "css transform: rotateY/X based on mouse position; banner art lives in card"},
    {"screen": "Encounter selection", "element": "Selected card pop-out", "notes": "Hovered card lifts + scales 1.15; siblings dim 70%"},
    {"screen": "Class selection", "element": "Class cards (rolling carousel)", "notes": "horizontal swipe / scroll; selected center is bigger"},
    {"screen": "Class selection", "element": "Class info panel", "notes": "Tagline, HP, kit summary, key cards preview"},
    {"screen": "Loading transition", "element": "Banner reveal", "notes": "Encounter banner sweeps in left→right; tip-of-the-game text below"},
    {"screen": "Victory / Defeat", "element": "Outcome card", "notes": "Slow zoom on monster portrait or party shot; round count + summary stats"},
]

SOUND = [
    {"slot": "ui-click",       "duration": "60-120ms", "format": "OGG/MP3", "notes": "soft click for any button"},
    {"slot": "ui-hover",       "duration": "40-80ms",  "format": "OGG/MP3", "notes": "softer than click; subtle"},
    {"slot": "card-play",      "duration": "200-400ms","format": "OGG/MP3", "notes": "per role: attack swoosh / defend low thud / heal chime"},
    {"slot": "monster-hit",    "duration": "300-600ms","format": "OGG/MP3", "notes": "impact thud + flesh"},
    {"slot": "monster-defeat", "duration": "1.0-1.5s", "format": "OGG/MP3", "notes": "shatter / dissolve"},
    {"slot": "round-resolve",  "duration": "1.0s",     "format": "OGG/MP3", "notes": "transition whoosh"},
    {"slot": "victory",        "duration": "2-3s",     "format": "OGG/MP3", "notes": "triumphant brass swell"},
    {"slot": "defeat",         "duration": "2-3s",     "format": "OGG/MP3", "notes": "low brass + reverb"},
    {"slot": "music-title",    "duration": "60-120s loop", "format": "OGG", "notes": "atmospheric drone"},
    {"slot": "music-combat",   "duration": "90-180s loop", "format": "OGG", "notes": "tension build per round"},
]


# ---------- WORKBOOK ----------

wb = Workbook()
ws_readme = wb.active
ws_readme.title = "README"

readme_rows = [
    ("Aetherfall — Master Asset Inventory",),
    ("",),
    ("Purpose",),
    ("This workbook lists every art / sound asset the game engine can consume, plus the dimensions, "
     "format, and naming conventions to use when generating or sourcing them. The companion file "
     "docs/asset-manifest.json carries the same data in a machine-readable format.",),
    ("",),
    ("Naming convention",),
    ("kebab-case file names rooted at assets/<category>/<sub>/<name>.<ext>",),
    ("Examples:",),
    ("  assets/cards/art/storm-forge/basic-attack.png",),
    ("  assets/ui/icons/hp.svg",),
    ("  assets/ui/banners/bruiser-duo.png    (already shipped)",),
    ("  assets/ui/effects/fx-damage.png      (sprite sheet)",),
    ("",),
    ("Format guidance",),
    ("  - Static UI icons: SVG preferred for scaling; PNG fallback OK",),
    ("  - Card art / banners / portraits: PNG, sRGB, no compression artifacts (the canvas is dark, "
     "so banding shows fast)",),
    ("  - Effect sprite sheets: PNG horizontal strip, frames 64x64 or 256x256, transparent BG",),
    ("  - Sound: OGG preferred; MP3 acceptable",),
    ("",),
    ("Sheet index",),
    ("  README — this sheet",),
    ("  Cards — every card that needs art (119 rows)",),
    ("  Classes — banner + portrait per class (9 rows)",),
    ("  Monsters — portrait per monster (8 rows)",),
    ("  Encounters — banner per encounter (5 rows)",),
    ("  Tokens — element-token icons (3 rows)",),
    ("  UI Icons — HP / energy / block / role / trait icons",),
    ("  UI Frames — panel chrome, baseplate, action bar, etc.",),
    ("  Effects — VFX sprite sheets",),
    ("  Backgrounds — battlefield + menu / loading / outcome",),
    ("  Sprite Sheets — preferred batch layouts",),
    ("  Menu / Title — AAA title screen + menu flow design spec",),
    ("  Sound — audio asset slots",),
    ("",),
    ("Update process",),
    ("  1. Drop a new image into the assets/ tree at the path listed in the relevant sheet.",),
    ("  2. Update src/content/game-assets.js so the engine references it.",),
    ("  3. Open a PR with a `feat: art —` commit prefix.",),
    ("",),
    ("Companion files",),
    ("  docs/asset-manifest.json — same inventory, machine-readable",),
    ("  docs/menu-design.md — deep-dive on the AAA title-screen / menu plan",),
]

for row in readme_rows:
    ws_readme.append(row)

ws_readme.column_dimensions["A"].width = 110
ws_readme.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="F5D083")
for r in (3, 6, 14, 20, 33):
    cell = ws_readme.cell(row=r, column=1)
    cell.font = Font(name="Calibri", size=11, bold=True, color="F5E1A8")

# --- Cards sheet ---
ws_cards = wb.create_sheet("Cards")
card_columns = [
    "Card ID", "Class", "Card Name", "Role", "Energy Cost",
    "Art file path", "Art size (px)", "Format", "Status", "Generation prompt seed",
]
card_rows = []
for c in CARDS:
    klass = c["id"].split(".")[0]
    name_slug = c["id"].split(".")[1].replace("_", "-")
    art_path = f"assets/cards/art/{klass}/{name_slug}.png"
    prompt = (
        f"{c['name']} — {klass.replace('-', ' ').title()} {c['role']}-class card art. "
        "Square 768x768 hero illustration, subject centered, transparent or matte background, "
        "matches the class palette (see Classes sheet). Painterly digital art with subtle rim light."
    )
    card_rows.append([
        c["id"], klass, c["name"], c["role"], c["cost"],
        art_path, "768x768", "PNG sRGB", "needs art", prompt,
    ])
write_sheet(ws_cards, card_columns, card_rows)
ws_cards.column_dimensions["A"].width = 32
ws_cards.column_dimensions["F"].width = 50
ws_cards.column_dimensions["J"].width = 70

# --- Classes sheet ---
ws_classes = wb.create_sheet("Classes")
class_columns = [
    "Class ID", "Display Name", "Short Name", "Role", "Palette",
    "Banner path", "Banner size (px)", "Banner status",
    "Portrait path", "Portrait size (px)", "Portrait status",
    "Banner generation prompt",
]
class_rows = []
for cd in CLASSES:
    banner_path = f"assets/ui/banners/{cd['id']}.png"
    portrait_path = f"assets/ui/portraits/{cd['id']}.png"
    prompt = (
        f"{cd['name']} hero banner. 1500x1024, 16:9-ish landscape. "
        f"Character {cd['shortName']} ({cd['role']}) on left third, full body, dramatic stance. "
        f"Title text \"{cd['shortName'].upper()} STARTER DECK\" on right third, metallic chrome treatment. "
        f"Faction palette: {cd['palette']}. Cinematic painterly digital art, dark moody background, "
        f"soft rim light on character, optional faction sigil top-right."
    )
    class_rows.append([
        cd["id"], cd["name"], cd["shortName"], cd["role"], cd["palette"],
        banner_path, "1500x1024", cd["banner_status"],
        portrait_path, "512x512 (head + chest)", cd["portrait_status"],
        prompt,
    ])
write_sheet(ws_classes, class_columns, class_rows)
ws_classes.column_dimensions["A"].width = 22
ws_classes.column_dimensions["L"].width = 80

# --- Monsters sheet ---
ws_mon = wb.create_sheet("Monsters")
mon_columns = [
    "Monster ID", "Display Name", "Role", "Encounter",
    "Portrait path", "Portrait size (px)", "Status", "Generation prompt seed",
]
mon_rows = []
for m in MONSTERS:
    portrait_path = f"assets/ui/monsters/{m['id']}.png"
    prompt = (
        f"{m['name']} — {m['role']}. 512x512 character portrait, head + upper body, "
        "dark moody atmosphere, faction-appropriate color palette, painterly digital art. "
        "Match the encounter banner's mood (see Encounters sheet)."
    )
    mon_rows.append([
        m["id"], m["name"], m["role"], m["encounter"],
        portrait_path, "512x512", "needs portrait" if m["id"] not in {"hollow-titan"} else "see /assets/ui/monsters/",
        prompt,
    ])
write_sheet(ws_mon, mon_columns, mon_rows)
ws_mon.column_dimensions["A"].width = 24
ws_mon.column_dimensions["H"].width = 70

# --- Encounters sheet ---
ws_enc = wb.create_sheet("Encounters")
enc_columns = [
    "Encounter ID", "Display Name", "Monsters", "Banner path", "Banner size (px)", "Status", "Generation prompt seed",
]
enc_rows = []
for e in ENCOUNTERS:
    banner_path = f"assets/ui/banners/{e['id']}.png"
    prompt = (
        f"{e['name']} encounter banner. 1500x1024 landscape. Group composition of [{e['monsters']}]. "
        "Title text in bold metallic chrome at top center. Dramatic backlight, cinematic, painterly digital art."
    )
    enc_rows.append([
        e["id"], e["name"], e["monsters"], banner_path, "1500x1024", e["banner_status"], prompt,
    ])
write_sheet(ws_enc, enc_columns, enc_rows)
ws_enc.column_dimensions["A"].width = 26
ws_enc.column_dimensions["G"].width = 80

# --- Tokens sheet ---
ws_tok = wb.create_sheet("Tokens")
tok_columns = ["Token ID", "Label", "Color", "Passive effect", "Icon path", "Icon size", "Generation prompt"]
tok_rows = []
for t in TOKENS:
    tok_rows.append([
        t["id"], t["label"], t["color"], t["passive"],
        f"assets/ui/icons/tokens/{t['id']}.png", t["icon_size"], t["prompt"],
    ])
write_sheet(ws_tok, tok_columns, tok_rows)
ws_tok.column_dimensions["G"].width = 90

# --- UI Icons sheet ---
ws_ui_icons = wb.create_sheet("UI Icons")
ui_icon_columns = ["Name", "Key", "Size", "Format", "Color guidance", "Engine usage"]
ui_icon_rows = [[i["name"], i["key"], i["size"], i["format"], i["color"], i["usage"]] for i in UI_ICONS]
write_sheet(ws_ui_icons, ui_icon_columns, ui_icon_rows)

# --- UI Frames sheet ---
ws_frames = wb.create_sheet("UI Frames")
frame_columns = ["Name", "Key", "Size", "Format", "Notes"]
frame_rows = [[f["name"], f["key"], f["size"], f["format"], f["notes"]] for f in UI_FRAMES]
write_sheet(ws_frames, frame_columns, frame_rows)
ws_frames.column_dimensions["E"].width = 60

# --- Effects sheet ---
ws_fx = wb.create_sheet("Effects")
fx_columns = ["Name", "Key", "Size", "Format", "Notes"]
fx_rows = [[f["name"], f["key"], f["size"], f["format"], f["notes"]] for f in EFFECTS]
write_sheet(ws_fx, fx_columns, fx_rows)
ws_fx.column_dimensions["E"].width = 60

# --- Backgrounds sheet ---
ws_bg = wb.create_sheet("Backgrounds")
bg_columns = ["Name", "Key", "Size", "Format", "Status"]
bg_rows = [[b["name"], b["key"], b["size"], b["format"], b["status"]] for b in BACKGROUNDS]
write_sheet(ws_bg, bg_columns, bg_rows)
ws_bg.column_dimensions["E"].width = 60

# --- Sprite sheets ---
ws_sheet = wb.create_sheet("Sprite Sheets")
sheet_columns = ["Name", "Key", "Layout", "Cell size", "Total size", "Contents"]
sheet_rows = [[s["name"], s["key"], s["layout"], s["cell"], s["total"], s["contains"]] for s in SPRITESHEETS]
write_sheet(ws_sheet, sheet_columns, sheet_rows)

# --- Menu / Title ---
ws_menu = wb.create_sheet("Menu and Title")
menu_columns = ["Screen", "Element", "Implementation notes"]
menu_rows = [[m["screen"], m["element"], m["notes"]] for m in MENU_SCREENS]
write_sheet(ws_menu, menu_columns, menu_rows)
ws_menu.column_dimensions["C"].width = 70

# --- Sound ---
ws_sound = wb.create_sheet("Sound")
sound_columns = ["Slot", "Duration", "Format", "Notes"]
sound_rows = [[s["slot"], s["duration"], s["format"], s["notes"]] for s in SOUND]
write_sheet(ws_sound, sound_columns, sound_rows)
ws_sound.column_dimensions["D"].width = 60

OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_XLSX)
print(f"Wrote {OUT_XLSX}")

# ---------- JSON manifest ----------

manifest = {
    "version": 1,
    "generated_at": "2026-05-10",
    "naming_convention": "kebab-case under assets/<category>/<sub>/<name>.<ext>",
    "image_format_guidance": {
        "icons": "SVG preferred, PNG fallback",
        "card_art": "PNG sRGB, 768x768 hero illustration, subject centered, alpha channel optional",
        "banners": "PNG, 1500x1024 landscape, sRGB",
        "portraits": "PNG, 512x512 head + upper body, sRGB",
        "effects": "PNG horizontal sprite-strip, transparent BG, frames sized 64x64 or 256x256",
        "backgrounds": "PNG/JPG 1920x1080+, sRGB",
    },
    "categories": {
        "cards": {
            "default_size_px": [768, 768],
            "format": "PNG",
            "path_template": "assets/cards/art/{class_id}/{card_slug}.png",
            "items": [
                {
                    "id": c["id"],
                    "class_id": c["id"].split(".")[0],
                    "name": c["name"],
                    "role": c["role"],
                    "cost": c["cost"],
                }
                for c in CARDS
            ],
        },
        "classes": {
            "banner": {"size_px": [1500, 1024], "format": "PNG", "path_template": "assets/ui/banners/{class_id}.png"},
            "portrait": {"size_px": [512, 512], "format": "PNG", "path_template": "assets/ui/portraits/{class_id}.png"},
            "items": [{"id": c["id"], "name": c["name"], "role": c["role"], "palette": c["palette"]} for c in CLASSES],
        },
        "monsters": {
            "portrait": {"size_px": [512, 512], "format": "PNG", "path_template": "assets/ui/monsters/{monster_id}.png"},
            "items": MONSTERS,
        },
        "encounters": {
            "banner": {"size_px": [1500, 1024], "format": "PNG", "path_template": "assets/ui/banners/{encounter_id}.png"},
            "items": ENCOUNTERS,
        },
        "tokens": {
            "icon": {"size_px": [64, 64], "format": "PNG transparent", "path_template": "assets/ui/icons/tokens/{token_id}.png"},
            "items": TOKENS,
        },
        "ui_icons": {
            "default_size_px": [32, 32],
            "format": "SVG preferred, PNG fallback",
            "path_template": "assets/ui/icons/{key}.svg",
            "items": UI_ICONS,
        },
        "ui_frames": {"items": UI_FRAMES},
        "effects": {"items": EFFECTS},
        "backgrounds": {"items": BACKGROUNDS},
        "sprite_sheets": {"items": SPRITESHEETS},
        "menu_screens": {"items": MENU_SCREENS},
        "sound": {"items": SOUND},
    },
}

with OUT_JSON.open("w") as f:
    json.dump(manifest, f, indent=2)
print(f"Wrote {OUT_JSON}")
